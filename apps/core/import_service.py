"""Bulk spreadsheet import: parsing, per-kind row processors, .xlsx templates,
and a non-blocking background runner that streams progress into BulkImport.

Supports .xlsx (openpyxl) and .csv. Each data row is validated and applied in
its own transaction so one bad row never aborts the batch; per-row errors are
recorded for display.
"""

import csv
import io
import logging
import re
import threading
import time
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone

from .models import BulkImport

logger = logging.getLogger(__name__)

MAX_ERRORS_KEPT = 500  # cap stored errors so a fully-bad file can't bloat the row


# --------------------------------------------------------------------------- #
# Templates (header columns + a sample row) per import kind
# --------------------------------------------------------------------------- #
TEMPLATES = {
    BulkImport.Kind.CUSTOMERS: {
        "columns": ["phone", "name", "email", "address", "latitude", "longitude"],
        "sample": ["+919876500001", "Asha Verma", "asha@example.com", "12 MG Road, Kanpur", "26.7422279", "82.142914"],
        # (column, required?, how to fill)
        "fields": [
            ("phone", "Required", "Mobile number. 10–15 digits; spaces/dashes are cleaned automatically. e.g. +919876500001 or 9876500001"),
            ("name", "Optional", "Customer's full name."),
            ("email", "Optional", "Email address."),
            ("address", "Optional", "Full delivery address. Saved as the customer's delivery address."),
            ("latitude", "Optional", "Decimal latitude for live tracking. e.g. 26.7422279"),
            ("longitude", "Optional", "Decimal longitude. e.g. 82.142914"),
        ],
        "notes": [
            "Existing customers (matched by phone) are updated, not duplicated.",
            "Extra columns (e.g. Segment, Enumerator, Quantity) are ignored.",
            "Header names are flexible: 'CONSUMER NAME', 'PHONE NO', 'ADDRESS (...latitude)' etc. are recognised.",
        ],
    },
    BulkImport.Kind.RIDERS: {
        "columns": ["phone", "name", "email", "address", "vehicle_number"],
        "sample": ["+919876500002", "Ravi Kumar", "ravi@example.com", "Faizabad 224001", "UP78 AB 1234"],
        "fields": [
            ("phone", "Required", "Rider's mobile number. 10–15 digits."),
            ("name", "Optional", "Rider's full name."),
            ("email", "Optional", "Email address."),
            ("address", "Optional", "Home/base address."),
            ("vehicle_number", "Optional", "Vehicle registration, e.g. UP78 AB 1234"),
        ],
        "notes": [
            "Creates a delivery-partner profile for the user.",
            "If the phone already belongs to a rider, that row is reported as an error.",
        ],
    },
    BulkImport.Kind.INVENTORY: {
        "columns": ["sku", "restock_quantity", "set_stock"],
        "sample": ["md-full-cream-milk-500-ml", "50", ""],
        "fields": [
            ("sku", "Required", "The exact SKU CODE of a variant (see the Inventory tab) — NOT the product name. e.g. md-full-cream-milk-500-ml"),
            ("restock_quantity", "Optional", "ADD this many units to current stock. e.g. 50"),
            ("set_stock", "Optional", "SET stock to this exact number. e.g. 100"),
        ],
        "notes": [
            "Provide EITHER restock_quantity OR set_stock per row.",
            "If both are filled, set_stock is used.",
            "Unknown SKUs are reported as errors — copy SKU codes from the Inventory tab.",
        ],
    },
}


def existing_rows(kind):
    """Current records for a kind, as row lists matching the template columns.

    Lets the downloaded template come pre-populated so the operator edits live
    data and re-uploads, instead of starting from a blank sheet.
    """
    if kind == BulkImport.Kind.CUSTOMERS:
        from apps.accounts.models import User

        rows = []
        for u in User.objects.filter(role=User.Role.CUSTOMER).prefetch_related("addresses")[:5000]:
            addr = u.addresses.first()  # default first (model ordering)
            lat = str(addr.latitude) if addr and addr.latitude is not None else ""
            lng = str(addr.longitude) if addr and addr.longitude is not None else ""
            line = (addr.address_line if addr else "") or u.address
            rows.append([u.phone, u.name, u.email, line, lat, lng])
        return rows
    if kind == BulkImport.Kind.RIDERS:
        from apps.delivery.models import DeliveryPartner

        return [
            [r.user.phone, r.user.name, r.user.email, r.user.address, r.vehicle_number]
            for r in DeliveryPartner.objects.select_related("user")[:5000]
        ]
    if kind == BulkImport.Kind.INVENTORY:
        from apps.catalog.models import ProductVariant

        return [
            [v.sku, "", v.stock]
            for v in ProductVariant.objects.order_by("product__name")[:5000]
        ]
    return []


def template_bytes(kind):
    """Return a two-sheet .xlsx: a data sheet (pre-filled with current records,
    or a sample row when empty) + an Instructions sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    spec = TEMPLATES[kind]
    wb = Workbook()

    # Sheet 1 — data (header row + existing records, or a sample row)
    ws = wb.active
    ws.title = kind.capitalize()[:31]
    ws.append(spec["columns"])
    header_fill = PatternFill("solid", fgColor="DEF9EC")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="253D4E")
        cell.fill = header_fill
    data = existing_rows(kind)
    if data:
        for row in data:
            ws.append(row)
    elif spec.get("sample"):
        ws.append(spec["sample"])
    for i in range(1, len(spec["columns"]) + 1):
        ws.column_dimensions[chr(64 + i)].width = 26

    # Sheet 2 — instructions
    info = wb.create_sheet("Instructions")
    title = info.cell(row=1, column=1, value="How to fill the “%s” sheet" % ws.title)
    title.font = Font(bold=True, size=14, color="253D4E")
    headers = ["Column", "Required?", "How to fill"]
    info.append([])
    info.append(headers)
    for cell in info[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3BB77E")
    for col, required, how in spec["fields"]:
        info.append([col, required, how])
    info.append([])
    notes_label = info.cell(row=info.max_row + 1, column=1, value="Notes")
    notes_label.font = Font(bold=True, color="253D4E")
    if data:
        info.append(["•", "This file is pre-filled with your current %d record(s) — edit values and re-upload to update." % len(data)])
    for note in spec.get("notes", []):
        info.append(["•", note])
    widths = {"A": 22, "B": 12, "C": 80}
    for col, width in widths.items():
        info.column_dimensions[col].width = width
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #
class ParseError(Exception):
    pass


def parse_spreadsheet(django_file):
    """Parse an uploaded .xlsx/.csv into a list of header-keyed dicts."""
    name = (django_file.name or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(django_file)
    if name.endswith(".xlsx"):
        return _parse_xlsx(django_file)
    raise ParseError("Unsupported file type. Upload a .xlsx or .csv file.")


def _clean(v):
    return "" if v is None else str(v).strip()


def _parse_csv(django_file):
    text = io.TextIOWrapper(django_file.file, encoding="utf-8-sig", newline="")
    reader = csv.DictReader(text)
    if not reader.fieldnames:
        raise ParseError("The file is empty.")
    rows = []
    for raw in reader:
        row = {(_clean(k)).lower(): _clean(v) for k, v in raw.items() if k}
        if any(row.values()):
            rows.append(row)
    return rows


def _parse_xlsx(django_file):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(django_file, read_only=True, data_only=True)
    except Exception:
        raise ParseError("Couldn't read the Excel file — is it a valid .xlsx?")
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = [_clean(h).lower() for h in next(it)]
    except StopIteration:
        raise ParseError("The file is empty.")
    if not any(header):
        raise ParseError("Missing a header row.")
    rows = []
    for raw in it:
        if raw is None or all(c is None or _clean(c) == "" for c in raw):
            continue
        row = {}
        for i, key in enumerate(header):
            if not key:
                continue
            row[key] = _clean(raw[i]) if i < len(raw) else ""
        rows.append(row)
    return rows


# --------------------------------------------------------------------------- #
# Flexible column mapping — tolerate real-world headers / spacing variations
# --------------------------------------------------------------------------- #
_PHONE_STRIP = re.compile(r"[^\d+]")

# alias → list of accepted (lowercased) header names
ALIASES = {
    "phone": ["phone", "phone no", "phone no.", "phone number", "mobile", "mobile no", "contact"],
    "name": ["name", "consumer name", "customer name", "full name", "consumer"],
    "email": ["email", "email id", "e-mail"],
    "address": ["address", "address line", "full address", "addr"],
    "latitude": ["latitude", "lat", "address (location answered) (latitude)"],
    "longitude": ["longitude", "lng", "lon", "address (location answered) (longitude)"],
    "vehicle_number": ["vehicle_number", "vehicle number", "vehicle", "vehicle no"],
    "sku": ["sku", "variant sku", "code"],
    "restock_quantity": ["restock_quantity", "restock quantity", "restock", "quantity", "qty"],
    "set_stock": ["set_stock", "set stock", "stock"],
    "city": ["city", "town"],
    "state": ["state"],
    "pincode": ["pincode", "pin", "zip", "postal code"],
}


def _pick(row, key):
    for header in ALIASES.get(key, [key]):
        if row.get(header, "") != "":
            return row[header]
    return ""


def _clean_phone(raw):
    raw = (raw or "").strip()
    plus = raw.startswith("+")
    digits = _PHONE_STRIP.sub("", raw).lstrip("+")
    return ("+" + digits) if plus else digits


def _to_decimal(value):
    try:
        return Decimal(str(round(float(value), 6)))
    except (ValueError, InvalidOperation, TypeError):
        return None


# --------------------------------------------------------------------------- #
# Per-kind row processors (raise ValueError with a human message on bad data)
# --------------------------------------------------------------------------- #
def _apply_profile(user, row):
    changed = []
    for field, key in (("name", "name"), ("email", "email"), ("address", "address")):
        value = _pick(row, key)
        if value:
            setattr(user, field, value)
            changed.append(field)
    if changed:
        user.save(update_fields=changed)


def _save_delivery_address(user, row):
    """Create/update a delivery address (with coordinates) from the row, if present."""
    from apps.addresses.models import Address

    line = _pick(row, "address")
    if not line:
        return
    defaults = {
        "city": _pick(row, "city"), "state": _pick(row, "state"), "pincode": _pick(row, "pincode"),
        "is_default": not user.addresses.exists(),
    }
    lat, lng = _to_decimal(_pick(row, "latitude")), _to_decimal(_pick(row, "longitude"))
    if lat is not None and lng is not None:
        defaults["latitude"], defaults["longitude"] = lat, lng
    Address.objects.update_or_create(user=user, address_line=line, defaults=defaults)


def _import_customer(row):
    from apps.accounts.models import User, phone_validator

    phone = _clean_phone(_pick(row, "phone"))
    if not phone:
        raise ValueError("phone is required")
    phone_validator(phone)
    user, created = User.objects.get_or_create(phone=phone)
    _apply_profile(user, row)
    _save_delivery_address(user, row)
    return created


def _import_rider(row):
    from apps.accounts.models import User, phone_validator
    from apps.delivery.models import DeliveryPartner

    phone = _clean_phone(_pick(row, "phone"))
    if not phone:
        raise ValueError("phone is required")
    phone_validator(phone)
    user, _ = User.objects.get_or_create(phone=phone)
    _apply_profile(user, row)
    if hasattr(user, "delivery_partner"):
        raise ValueError("already a rider")
    DeliveryPartner.objects.create(user=user, vehicle_number=_pick(row, "vehicle_number"))
    return True


def _import_inventory(row):
    from apps.catalog.models import ProductVariant
    from apps.inventory.models import StockMovement
    from apps.inventory.services import adjust_stock, restock

    sku = _pick(row, "sku")
    if not sku:
        raise ValueError("sku is required")
    try:
        variant = ProductVariant.objects.get(sku=sku)
    except ProductVariant.DoesNotExist:
        raise ValueError(f"unknown sku '{sku}' — use the SKU code from the Inventory tab, not the product name")

    set_stock = _pick(row, "set_stock")
    restock_qty = _pick(row, "restock_quantity")

    if set_stock != "":
        try:
            target = int(float(set_stock))
        except ValueError:
            raise ValueError(f"set_stock '{set_stock}' is not a number")
        if target < 0:
            raise ValueError("set_stock cannot be negative")
        delta = target - variant.stock
        if delta != 0:
            adjust_stock(variant, delta, StockMovement.Reason.ADJUSTMENT, note="Bulk import: set stock")
        return False
    if restock_qty != "":
        try:
            qty = int(float(restock_qty))
        except ValueError:
            raise ValueError(f"restock_quantity '{restock_qty}' is not a number")
        if qty <= 0:
            raise ValueError("restock_quantity must be positive")
        restock(variant, qty, note="Bulk import: restock")
        return False
    raise ValueError("provide either restock_quantity or set_stock")


PROCESSORS = {
    BulkImport.Kind.CUSTOMERS: _import_customer,
    BulkImport.Kind.RIDERS: _import_rider,
    BulkImport.Kind.INVENTORY: _import_inventory,
}


# --------------------------------------------------------------------------- #
# Runner (background thread; genuinely non-blocking in dev and prod)
# --------------------------------------------------------------------------- #
def _run(import_id, rows):
    job = BulkImport.objects.filter(pk=import_id).first()
    if not job:
        return
    processor = PROCESSORS[job.kind]
    BulkImport.objects.filter(pk=import_id).update(
        status=BulkImport.Status.PROCESSING, total_rows=len(rows), updated_at=timezone.now()
    )

    success = errors_count = 0
    errors = []
    pace = 0.04 if settings.DEBUG else 0  # in dev, pace slightly so progress is visible

    for index, row in enumerate(rows, start=1):
        try:
            with transaction.atomic():
                processor(row)
            success += 1
        except Exception as exc:  # noqa: BLE001 — one bad row must not stop the batch
            errors_count += 1
            if len(errors) < MAX_ERRORS_KEPT:
                errors.append({"row": index + 1, "message": str(exc)[:300]})  # +1: header row
        if pace:
            time.sleep(pace)
        if index % 5 == 0 or index == len(rows):
            BulkImport.objects.filter(pk=import_id).update(
                processed_rows=index, success_count=success, error_count=errors_count,
                errors=errors, updated_at=timezone.now(),
            )

    BulkImport.objects.filter(pk=import_id).update(
        status=BulkImport.Status.COMPLETED, processed_rows=len(rows),
        success_count=success, error_count=errors_count, errors=errors, updated_at=timezone.now(),
    )


def start_import(import_id, rows):
    """Kick off processing in a background daemon thread (non-blocking)."""
    def worker():
        try:
            _run(import_id, rows)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Bulk import %s failed", import_id)
            BulkImport.objects.filter(pk=import_id).update(
                status=BulkImport.Status.FAILED, message=str(exc)[:255], updated_at=timezone.now()
            )
        finally:
            connection.close()  # release this thread's DB connection

    threading.Thread(target=worker, daemon=True).start()
